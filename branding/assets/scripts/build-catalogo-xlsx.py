"""
Genera branding/catalogo-productos.xlsx desde el catálogo redactado en
branding/catalogo-productos.md.

Embebe miniaturas de la primera foto de cada producto en una columna "Foto",
para que Felipe pueda revisar visualmente sin abrir archivos por separado.

Uso:
    cd /root/lasercrafts
    python3 branding/assets/scripts/build-catalogo-xlsx.py

Salidas:
    branding/catalogo-productos.xlsx
    branding/assets/productos/thumbs/*.jpg   (miniaturas embebidas)

Requiere:
    - openpyxl
    - Pillow
    - Las fotos crudas en branding/assets/productos/fotos/ (Drive de Felipe,
      no van al repo). Si no están, los thumbs no se generan y el Excel
      sale sin imágenes (con un "—" en la columna Foto).
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from PIL import Image, ImageOps

# --- Paths ------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[3]
FOTOS_DIR = ROOT / "branding/assets/productos/fotos"
THUMBS_DIR = ROOT / "branding/assets/productos/thumbs"
OUT_XLSX = ROOT / "branding/catalogo-productos.xlsx"

THUMBS_DIR.mkdir(parents=True, exist_ok=True)

# --- Paleta v0.3 ------------------------------------------------------------
PROFUNDO = "0D4080"
NOGAL = "703D24"
PAPEL = "FFFFFF"
NOGAL_12 = "F2EBE6"

# --- Productos --------------------------------------------------------------
# (slug, Categoría, Título, Bajada, Descripción, Materialidad, Personalización,
#  Variantes, [lista de fotos — la primera se usa como thumb])
PRODUCTS = [
    ("posavasos-valdivia",
     "Posavasos y mesa",
     "Posavasos Valdivia",
     "Posavasos cuadrado en terciado con un pájaro geométrico y la palabra Valdivia.",
     "Lo grabamos cuadrado, con un pájaro de líneas geométricas y \"Valdivia\" en cursiva al lado. Atrás lleva imán, así también sirve de souvenir de refrigerador. Sale solo o en sets de 4 y 6.",
     "Terciado natural, grabado láser por una cara. Reverso liso con imán.",
     "Cambiar la palabra (otra ciudad, un nombre, una fecha).",
     "Suelto / Set 4 / Set 6",
     ["20260328_123630(1).jpg", "20260328_123628.jpg"]),

    ("posavasos-fauna-nativa",
     "Posavasos y mesa",
     "Posavasos fauna nativa",
     "Posavasos redondos con grabado de fauna nativa y su nombre científico.",
     "Cada pieza es una especie del sur de Chile, grabada en línea fina con nombre común y científico al borde. Hoy hemos hecho cisne cuello negro, martín pescador y otras aves del bosque valdiviano. Se piden por especie o como set temático (parques, viñas, hoteles).",
     "Terciado, grabado por una cara, sin pintar.",
     "Elegir especies, agregar logo institucional al reverso.",
     "Cisne cuello negro / Martín pescador / a definir",
     ["20260531_135930.jpg", "20260531_140635.jpg"]),

    ("set-posavasos-geometricos",
     "Posavasos y mesa",
     "Set posavasos geométricos + cofre",
     "Cuatro posavasos calados con patrones geométricos, presentados en cofre con la firma LáserCrafts.",
     "Cuatro patrones distintos —asanoha, seigaiha y dos más— calados en terciado, listos para servir en mesa de café. Vienen en un cofre cuadrado con la firma del taller grabada al fondo. Es nuestro regalo corporativo más pedido.",
     "Terciado, calado pasante, sin pintar.",
     "Grabado en la tapa o el frente del cofre con el logo o nombre del cliente.",
     "Set 4 patrones (estándar) / personalizado",
     ["20260328_122332.jpg", "20260328_122324.jpg"]),

    ("bandeja-portavasos",
     "Posavasos y mesa",
     "Bandeja porta-posavasos",
     "Bandeja baja con la firma LáserCrafts grabada al frente, lista para una mesa de café.",
     "La bandeja sostiene seis posavasos y una taza al medio. La grabamos al frente con el nombre de quien la regala o la marca del local que la pide. Cabe en mesa de centro o repisa de cocina.",
     "MDF con grabado láser por una cara.",
     "Texto al frente (nombre, marca, frase).",
     "Estándar / personalizado",
     ["20260328_115416.jpg"]),

    ("cofre-tapa-hojas",
     "Cofres y cajas",
     "Cofre tapa de hojas",
     "Cofre rectangular con tapa de hojas caladas, para guardar lo que sea importante.",
     "La tapa es una composición de hojas caladas que se cruzan; la luz pasa entre ellas. Adentro cabe joyería, un anillo de matrimonio, recuerdos. Es el cofre que más pedimos para nacimientos y aniversarios.",
     "Terciado, calado pasante en la tapa, cuerpo cerrado.",
     "Grabado interior (nombre, fecha) en el fondo de la caja.",
     "Estándar / con grabado interior",
     ["20260328_110045.jpg", "20260328_110159.jpg"]),

    ("cofre-cuadrado-firma",
     "Cofres y cajas",
     "Cofre cuadrado firma LáserCrafts",
     "Cofre cuadrado con la firma del taller grabada al fondo.",
     "Lo armamos a la medida del set de posavasos, pero también funciona como cofre solo. La firma queda escondida al fondo, así el cliente la descubre al abrir.",
     "Terciado, grabado al fondo.",
     "Mensaje grabado al fondo (sustituye la firma) o en la tapa.",
     "Solo / con set de posavasos",
     ["20260328_122332.jpg", "20260328_122324.jpg"]),

    ("soporte-celular-plegable",
     "Soportes para escritorio",
     "Soporte de celular plegable",
     "Soporte de celular que se pliega, con grabado al frente.",
     "Lo grabamos con el logo de un equipo, una marca o una frase. Se pliega cuando no se usa y queda como pieza de escritorio. Sirve para celular y para tablet pequeña.",
     "MDF, grabado láser al frente.",
     "Grabado al frente (escudo de equipo, logo, frase, nombre).",
     "Logo personalizado / estándar",
     ["20260328_123021.jpg"]),

    ("soporte-celular-engranaje",
     "Soportes para escritorio",
     "Soporte de celular engranaje",
     "Soporte para celular con engranaje calado y madera oscura.",
     "El detalle del engranaje queda visible al apoyar el celular. Lo entregamos en madera oscura para que el contraste con el grabado sea fuerte. Es nuestro soporte más vendido a personas que trabajan en escritorio.",
     "MDF teñido oscuro, calado y armado por encastre.",
     "Grabado al lateral del engranaje.",
     "Estándar / con grabado lateral",
     ["20260328_162735.jpg"]),

    ("pins-imanes-naturaleza",
     "Pins, imanes y llaveros",
     "Set pins/imanes naturaleza",
     "Set de cuatro pins pintados a mano: hongo amanita, pino con luna, hongo violeta, mariposa luna.",
     "Cada pieza la pintamos a mano sobre el grabado, así no hay dos exactamente iguales. Se piden sueltos o en sets de cuatro. Se usan como pin de mochila, imán de refrigerador o adorno de planta.",
     "Terciado o MDF, grabado láser, pintado a mano.",
     "Elegir las cuatro piezas del set entre las disponibles.",
     "Suelto / Set 4",
     ["20260328_111352.jpg", "20260328_111414.jpg", "20260328_111253.jpg"]),

    ("llaveros-fauna-nativa",
     "Pins, imanes y llaveros",
     "Llaveros fauna nativa",
     "Llaveros de pájaros y mamíferos del bosque valdiviano, con nombre al lado.",
     "Hoy tenemos chucao y monito del monte; vamos sumando especies según pedidos. El reverso queda limpio para grabar el logo de quien los regala (parque, hotel, viña). Llevan anilla metálica y cadena corta.",
     "Terciado, grabado láser por las dos caras, anilla metálica.",
     "Logo institucional al reverso, set por temática (aves, mamíferos, hongos).",
     "Chucao / Monito del monte / + a definir",
     ["20260531_130739.jpg", "20260531_131826.jpg", "20260531_131635.jpg"]),

    ("marcapaginas-fauna",
     "Marcapáginas",
     "Marcapáginas fauna nativa",
     "Marcapáginas en terciado con flora y hongos del sur, nombre común y científico.",
     "La forma sigue al organismo —el digüeñe es una cabeza de racimos, el campañita magallánica es un hongo entero con su raíz, otros son árboles con copa y raíz. En el cabezal va el nombre común y el científico abajo. Pensados para librería independiente, museo o regalo botánico.",
     "Terciado, grabado por una cara, contorno calado.",
     "Elegir las especies del set, agregar logo en el cabezal.",
     "Digüeñe / Campañita magallánica / + a definir",
     ["20260531_130023.jpg", "20260531_125703.jpg", "20260531_130419.jpg"]),

    ("cuadro-3d-pez-betta",
     "Cuadros 3D multicapa",
     "Cuadro 3D pez betta",
     "Cuadro enmarcado con un pez betta en relieve sobre fondo de peces grabados.",
     "El pez se arma en varias capas de madera teñida, así sus aletas tienen profundidad real (no es solo grabado). El fondo es una trama de peces más pequeños grabados en plano. Marco oscuro, listo para colgar.",
     "Terciado multicapa, teñido, marco oscuro.",
     "Elegir tonos del pez (rojo, azul, dorado).",
     "Rojo / Azul / Dorado",
     ["20260328_111942.jpg"]),

    ("cuadro-3d-manga",
     "Cuadros 3D multicapa",
     "Cuadro 3D personaje manga/anime",
     "Retrato 3D armado en capas: personaje sobre fondo decorativo, marco listo para colgar.",
     "Lo hacemos a pedido sobre un personaje que mande el cliente. Se arma en cuatro a seis capas, con colores planos en rojo, azul, gris y negro. Llega con bastidor, listo para colgar.",
     "Terciado multicapa, teñido a mano, montado en bastidor.",
     "A medida — el cliente manda el personaje o la imagen base.",
     "A medida",
     ["20260328_113614.jpg"]),

    ("carta-pokemon-3d",
     "Coleccionables Pokémon",
     "Carta Pokémon 3D",
     "Reproducción 3D de carta Pokémon armada en capas y pintada a mano.",
     "El holograma de la carta se reemplaza por capas de madera grabadas y pintadas: el Pokémon en relieve, el fondo, el contorno y el texto. Cada una toma tiempo, así que las hacemos a pedido. Hoy hemos hecho Leafeon Vmax, Pikachu y Wartortle, pero se puede pedir otra.",
     "Terciado multicapa, grabado y pintado a mano.",
     "A medida — el cliente elige el Pokémon, la carta y los colores.",
     "Leafeon Vmax / Pikachu / Wartortle / + a pedido",
     ["20260328_123148.jpg", "20260328_123132.jpg", "20260328_122951.jpg"]),

    ("medallas-pokemon",
     "Coleccionables Pokémon",
     "Set medallas de gimnasio",
     "Set de seis medallas de gimnasio pintadas a mano.",
     "Reproducimos las medallas clásicas: roca, agua, fuego, hoja, espina, planta, alma y otras. Se piden por unidad o como set completo, atrás llevan imán. Es regalo cumpleaños fan #1 del año.",
     "MDF, grabado, pintado a mano, imán al reverso.",
     "Elegir qué medallas entran en el set.",
     "Suelto / Set 6 / Set 8",
     ["20260328_111549.jpg"]),

    ("casa-halloween-armable",
     "Casitas armables y maquetas",
     "Casa Halloween armable",
     "Casa fantasma armable en madera, con espacio para vela LED y figuras alrededor.",
     "Llega plana, se arma por encastre. Lleva fachada con telarañas grabadas, ventanas troqueladas, calabaza al frente, fantasma colgando y carteles de \"Ghost Crossing\" y \"Witches Way\". El espacio interior recibe una vela LED para que las ventanas brillen.",
     "MDF, pintado a mano, encastre.",
     "Nombre del niño/a en el frente, año de la fiesta.",
     "Estándar / personalizada",
     ["20260328_114526.jpg"]),

    ("placa-direccion-escena",
     "Placas y letreros",
     "Placa de dirección con escena",
     "Placa para la entrada con número de calle y una escena grabada del lugar.",
     "El cliente manda foto de su casa, dibujo o referencia; nosotros la traducimos a grabado. Va con el número de calle grabado grande y la escena al lado. Se cuelga afuera o en el living como recuerdo del lugar.",
     "Terciado, grabado por una cara, perforaciones para colgar.",
     "A medida — el dibujo, el número, el nombre de la familia.",
     "A medida",
     ["20260328_120002.jpg"]),

    ("letrero-exterior-tienda",
     "Placas y letreros",
     "Letrero exterior para tienda",
     "Letrero colgante en madera con grabado del logo y tagline de la marca.",
     "Lo hacemos para comercios pequeños que quieren un cartel exterior que no parezca impreso. Trabajamos con vectorial del cliente o lo redibujamos juntos. Va terminado con barniz exterior y herrajes para colgar.",
     "Terciado o tablón, grabado, pintado, barniz exterior, herraje de hierro.",
     "A medida — el logo, los colores, las dimensiones.",
     "A medida",
     ["20260328_115809.jpg"]),

    ("expositor-hot-wheels",
     "Encargos a medida",
     "Expositor Hot Wheels",
     "Expositor mural en MDF para colección de Hot Wheels con logo grabado.",
     "Lo armamos para coleccionistas serios: tres columnas, soportes para que las cajas queden inclinadas y visibles, logo Hot Wheels grabado al frente. Cabe en muro y se llena de a poco.",
     "MDF teñido oscuro, grabado, herraje para muro.",
     "A medida — cantidad de columnas, tipo de colección (Pokémon, Funko, Tarot, etc.).",
     "A medida",
     ["IMG-20260328-WA0058.jpg"]),
]

# --- Generar miniaturas ----------------------------------------------------
THUMB_W = 280  # ancho final del thumb en px (suficiente para celda Excel ~38u)

def make_thumb(slug: str, photo_name: str) -> Path | None:
    """Genera el thumb si la foto fuente existe. Devuelve la ruta del thumb."""
    src = FOTOS_DIR / photo_name
    if not src.exists():
        return None
    dst = THUMBS_DIR / f"{slug}.jpg"
    if dst.exists():
        return dst
    with Image.open(src) as im:
        im = ImageOps.exif_transpose(im)        # honrar rotación EXIF
        im.thumbnail((THUMB_W, THUMB_W * 2))    # cap por ancho, alto libre
        im = im.convert("RGB")
        im.save(dst, "JPEG", quality=78, optimize=True)
    return dst

thumbs_generados = 0
thumbs_faltantes = []
for p in PRODUCTS:
    slug, *_rest, fotos = p
    first = fotos[0] if fotos else None
    if not first:
        thumbs_faltantes.append((slug, "(sin foto referenciada)"))
        continue
    dst = make_thumb(slug, first)
    if dst is None:
        thumbs_faltantes.append((slug, first))
    else:
        thumbs_generados += 1

print(f"Thumbs generados: {thumbs_generados}/{len(PRODUCTS)}")
if thumbs_faltantes:
    print("Sin thumb (foto fuente no encontrada):")
    for slug, fn in thumbs_faltantes:
        print(f"  - {slug}: {fn}")

# --- Construcción del workbook --------------------------------------------
wb = Workbook()

# === Hoja 1: Productos ====================================================
ws = wb.active
ws.title = "Productos"
ws.sheet_view.zoomScale = 100
ws.sheet_view.showGridLines = False

# Encabezado de marca
ws.merge_cells("A1:O1")
c = ws["A1"]
c.value = "LáserCrafts · Catálogo de productos · v0.1"
c.font = Font(name="Inter", bold=True, size=16, color=PAPEL)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
c.fill = PatternFill("solid", fgColor=PROFUNDO)
ws.row_dimensions[1].height = 38

# Sub-instrucciones
ws.merge_cells("A2:O2")
c = ws["A2"]
c.value = ("Edita las columnas grises (dimensiones, plazo, precio, foto definitiva, visible, orden). "
          "Las azules son la propuesta del taller (título, bajada, descripción) — modifícalas si no calzan con tu voz. "
          "Detalle de uso y voz de marca en la hoja \"Instrucciones\".")
c.font = Font(name="Inter", italic=True, size=10, color="3F3F3F")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
c.fill = PatternFill("solid", fgColor=NOGAL_12)
ws.row_dimensions[2].height = 38

# Headers de columna
# kind: ref / edit_blue / edit_gray
HEADERS = [
    ("Foto",                 16, "ref"),         # A — thumbnail
    ("Categoría",            22, "ref"),         # B
    ("Título",               32, "edit_blue"),   # C
    ("Bajada (lista)",       52, "edit_blue"),   # D
    ("Descripción (ficha)",  60, "edit_blue"),   # E
    ("Materialidad",         46, "edit_blue"),   # F
    ("Personalización",      46, "edit_blue"),   # G
    ("Variantes sugeridas",  34, "edit_blue"),   # H
    ("Foto referencia",      32, "ref"),         # I
    ("Dimensiones (mm)",     20, "edit_gray"),   # J
    ("Plazo (días)",         12, "edit_gray"),   # K
    ("Precio CLP",           14, "edit_gray"),   # L
    ("Foto definitiva",      24, "edit_gray"),   # M
    ("Mostrar en sitio (sí/no)", 20, "edit_gray"),# N
    ("Orden",                10, "edit_gray"),   # O
]

HEAD_ROW = 3
for col_idx, (name, width, _kind) in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEAD_ROW, column=col_idx, value=name)
    cell.font = Font(name="Inter", bold=True, size=11, color=PAPEL)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    cell.fill = PatternFill("solid", fgColor=PROFUNDO)
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[HEAD_ROW].height = 32

# Estilos por celda
thin = Side(border_style="thin", color="DDDDDD")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

BLUE_TINT = "E8EEF7"
GRAY_TINT = "F4F4F4"
REF_TINT = "FFFFFF"
KIND_FILL = {
    "ref":       PatternFill("solid", fgColor=REF_TINT),
    "edit_blue": PatternFill("solid", fgColor=BLUE_TINT),
    "edit_gray": PatternFill("solid", fgColor=GRAY_TINT),
}

# Alto de fila para que el thumb (~150px = 113pt) quepa
ROW_H = 150

for row_idx, p in enumerate(PRODUCTS, start=HEAD_ROW + 1):
    slug, categoria, titulo, bajada, descripcion, material, personal, variantes, fotos = p
    row_data = [
        "",                       # A: foto (se inserta como image, no texto)
        categoria,                # B
        titulo,                   # C
        bajada,                   # D
        descripcion,              # E
        material,                 # F
        personal,                 # G
        variantes,                # H
        ", ".join(fotos),         # I
        "",                       # J: Dimensiones
        "",                       # K: Plazo
        "",                       # L: Precio
        "",                       # M: Foto definitiva
        "",                       # N: Visible
        row_idx - HEAD_ROW,       # O: Orden
    ]
    for col_idx, (val, header) in enumerate(zip(row_data, HEADERS), start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name="Inter", size=10, color="1D1D1D")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        cell.fill = KIND_FILL[header[2]]
        cell.border = border
    ws.row_dimensions[row_idx].height = ROW_H

    # Insertar el thumb en columna A
    thumb_path = THUMBS_DIR / f"{slug}.jpg"
    if thumb_path.exists():
        img = XlImage(str(thumb_path))
        # Escalar para que el alto no exceda ~140px y el ancho ~120px
        max_h = 180
        max_w = 130
        ratio = min(max_w / img.width, max_h / img.height, 1.0)
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
        anchor_cell = f"A{row_idx}"
        ws.add_image(img, anchor_cell)
    else:
        # marcar "—" si no se pudo generar el thumb
        c = ws.cell(row=row_idx, column=1, value="—")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Inter", size=14, color="999999")

# Freeze panes: arriba (header) + columnas A (foto) + B (categoría) + C (título)
ws.freeze_panes = "D4"

# Autofilter
last_row = HEAD_ROW + len(PRODUCTS)
ws.auto_filter.ref = f"B{HEAD_ROW}:{get_column_letter(len(HEADERS))}{last_row}"

# === Hoja 2: Instrucciones ================================================
ws2 = wb.create_sheet("Instrucciones")
ws2.sheet_view.showGridLines = False

INSTR = [
    ("Cómo usar este archivo", "h1"),
    ("", "spacer"),
    ("Este Excel es el catálogo en curso de LáserCrafts. La primera hoja (Productos) tiene 19 productos identificados desde las fotos del taller. La idea es que Felipe revise cada fila y complete lo que falta.", "p"),
    ("", "spacer"),
    ("Columna Foto (A)", "h2"),
    ("Cada fila lleva una miniatura de la foto del taller para que sepas a qué producto nos referimos. Si quieres ver la foto original o las demás fotos del producto, mira la columna \"Foto referencia\" (I).", "p"),
    ("", "spacer"),
    ("Columnas en azul claro — propuesta del taller (texto)", "h2"),
    ("Título, bajada, descripción, materialidad, personalización y variantes vienen propuestos por defecto. Si la voz no te calza, edita el texto — la regla está en branding/voz-de-marca.md (registro \"taller arquitectónico\"): primera persona, frase corta, una idea por frase, cifras y fechas cuando aplique.", "p"),
    ("", "spacer"),
    ("Columnas en gris — datos del cliente", "h2"),
    ("Dimensiones (mm): alto × ancho × profundidad cuando aplique.", "li"),
    ("Plazo (días): tiempo real de fabricación desde confirmar el pedido.", "li"),
    ("Precio CLP: precio público B2C. Si tiene variantes, anota el rango (ej. \"19.000–29.000\").", "li"),
    ("Foto definitiva: nombre de la foto sobre fondo blanco/Nogal que vamos a usar en la ficha. Por ahora deja vacío o pon \"pendiente\".", "li"),
    ("Mostrar en sitio (sí/no): si es \"no\", queda fuera del catálogo público. Útil para productos que solo se hacen a pedido cerrado.", "li"),
    ("Orden: número que define la posición en el listado. 1 = primero. Felipe puede reordenar como quiera; el sitio lo respeta.", "li"),
    ("", "spacer"),
    ("Columnas en blanco — solo referencia", "h2"),
    ("Foto (A), Categoría (B) y Foto referencia (I) son para que sepas de qué producto hablamos. No es necesario editarlas, pero puedes si crees que el producto va en otra categoría.", "p"),
    ("", "spacer"),
    ("Cuando termines", "h2"),
    ("Manda este archivo de vuelta. Lo importo a web/maqueta/src/data/products.ts y aparece en lasercrafts.cl con las descripciones que tú aprobaste.", "p"),
    ("", "spacer"),
    ("Voz de marca — recordatorio rápido", "h2"),
    ("Sí: \"Lo grabo el martes\", \"$19.000\", \"antes del viernes\", terciado, taller, Felipe, yo, te.", "li"),
    ("No: \"calidad premium\", \"plazos competitivos\", \"nuestro equipo\", emojis dulces.", "li"),
    ("Detalle completo en branding/voz-de-marca.md.", "li"),
    ("", "spacer"),
    ("Pendientes generales del catálogo (§13 del .md)", "h2"),
    ("¿Productos Pokémon/manga se publican o solo a pedido? (IP de terceros)", "li"),
    ("¿El cofre porta-posavasos y el cofre cuadrado son uno o dos SKUs distintos?", "li"),
    ("¿Falta categoría \"Matrimonio\" (libro de firmas, plano de mesas, distintivos)?", "li"),
    ("¿Falta categoría \"Retratos del cliente\" (foto grabada en madera)?", "li"),
]

ws2.column_dimensions["A"].width = 110
row = 1
for text, kind in INSTR:
    cell = ws2.cell(row=row, column=1, value=text)
    if kind == "h1":
        cell.font = Font(name="Inter", bold=True, size=18, color=PROFUNDO)
        ws2.row_dimensions[row].height = 32
    elif kind == "h2":
        cell.font = Font(name="Inter", bold=True, size=12, color=NOGAL)
        ws2.row_dimensions[row].height = 22
    elif kind == "p":
        cell.font = Font(name="Inter", size=11, color="1D1D1D")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[row].height = 48
    elif kind == "li":
        cell.value = "·  " + text
        cell.font = Font(name="Inter", size=11, color="1D1D1D")
        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws2.row_dimensions[row].height = 26
    elif kind == "spacer":
        ws2.row_dimensions[row].height = 10
    row += 1

# Instrucciones primero
wb.move_sheet("Instrucciones", offset=-1)

# === Guardar ============================================================
wb.save(OUT_XLSX)
print(f"OK: {OUT_XLSX}  ({len(PRODUCTS)} productos)")
